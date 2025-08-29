import os
import time
import pandas as pd
from openpyxl import Workbook


class PkSimAnalyzer:
    def __init__(
        self,
        source_dir: str,
        output_dir: str,
        filter_params: list[tuple[str, str, bool, float]],
        mic_params: list[float],
        float_precision: int = 2
    ):
        self.source_dir = source_dir
        self.output_dir = output_dir
        self.filter_params = filter_params
        self.mic_params = mic_params
        self.float_precision = float_precision
        os.makedirs(self.output_dir, exist_ok=True)

    # ---------------- Core processing ---------------- #

    def process_all(self):
        for file_name in os.listdir(self.source_dir):
            if file_name.lower().endswith(".csv"):
                self.process_file(os.path.join(self.source_dir, file_name))

    def process_file(self, file_path: str):
        file_name = os.path.basename(file_path)
        base_name = os.path.splitext(file_name)[0]

        df = pd.read_csv(file_path)
        output_path = os.path.join(self.output_dir, f"{base_name}.xlsx")
        wb = Workbook()

        ws_main = wb.active
        ws_main.title = self._sanitize_sheet_name(base_name)
        self._write_dataframe_to_sheet(ws_main, df)

        for sheet_name, parameter_value, calculate_mic, conv_factor in self.filter_params:
            filtered_df = df[df["Parameter"] == parameter_value].copy()

            filtered_df = self._add_unit_conversion(filtered_df, conv_factor)

            if calculate_mic:
                filtered_df = self._add_mic_columns(filtered_df)

            ws = wb.create_sheet(title=self._sanitize_sheet_name(sheet_name))
            self._write_dataframe_to_sheet(ws, filtered_df)

            if calculate_mic:
                if parameter_value == "C_max_tD1_tD2":
                    self._add_pta_analysis(
                        wb,
                        filtered_df,
                        [
                            ("PTA Cmax ≥ 10", 10)
                        ],
                        use_mic=True)
                elif parameter_value == "AUC_tD1_tD2":
                    self._add_pta_analysis(
                        wb,
                        filtered_df,
                        [
                            ("PTA AUC ≥ 50", 50),
                            ("PTA AUC ≥ 110", 110),
                            ("PTA AUC ≥ 700", 700)
                        ],
                        use_mic=True)
            else:
                if parameter_value == "C_trough_tD2":
                    self._add_pta_analysis(wb, filtered_df, [("PTA Cmin ≥ 0.5", 0.5)], use_mic=False)

        wb.save(output_path)
        print(f"Processed and saved: {output_path}")

    # ---------------- Data transformations ---------------- #

    def _add_unit_conversion(self, df: pd.DataFrame, factor: float) -> pd.DataFrame:
        df = df.copy()
        df["Value"] = pd.to_numeric(df["Value"], errors="coerce")
        df["unit (ug.ml - mg.L)"] = (df["Value"] * factor).round(self.float_precision)
        return df

    def _add_mic_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        if "unit (ug.ml - mg.L)" not in df.columns:
            raise ValueError("Unit conversion must be applied before MIC calculation.")
        for mic in self.mic_params:
            col_name = f"MIC {mic} (ug.mL)"
            df[col_name] = (df["unit (ug.ml - mg.L)"] / mic).round(self.float_precision)
        return df

    # ---------------- PTA logic ---------------- #

    def _add_pta_analysis(self, wb, df: pd.DataFrame, thresholds: list[tuple[str, float]], use_mic: bool) -> None:
        """
        Add PTA analysis sheets.
        :param wb: Workbook
        :param df: DataFrame with MIC columns if use_mic=True
        :param thresholds: list of (label, threshold) tuples
        :param use_mic: True = threshold applied to MIC columns, False = applied to 'unit (ug.ml - mg.L)'
        """
        for label, threshold in thresholds:
            ws = wb.create_sheet(title=self._sanitize_sheet_name(label))

            if use_mic:
                ws.append(["MIC"] + self.mic_params)
                counts = []
                for mic in self.mic_params:
                    col = f"MIC {mic} (ug.mL)"
                    counts.append(int((df[col] >= threshold).sum()) if col in df.columns else 0)
                ws.append(["Count"] + counts)

                total = len(df)
                percentages = [round(c / total * 100, self.float_precision) if total > 0 else 0 for c in counts]
                ws.append(["%"] + percentages)

            else:
                ws.append(["Metric", "Value"])
                count_val = int((df["unit (ug.ml - mg.L)"] >= threshold).sum())
                ws.append(["Count", count_val])

                total = len(df)
                perc_val = round(count_val / total * 100, self.float_precision) if total > 0 else 0
                ws.append(["%", perc_val])

    # ---------------- Excel writing helpers ---------------- #

    def _sanitize_sheet_name(self, name: str) -> str:
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for ch in invalid_chars:
            name = name.replace(ch, "_")
        return name[:31]

    def _write_dataframe_to_sheet(self, sheet, df: pd.DataFrame) -> None:
        for col_idx, col_name in enumerate(df.columns, start=1):
            sheet.cell(row=1, column=col_idx, value=str(col_name))
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                if pd.isna(value):
                    safe_value = None
                elif hasattr(value, "item"):
                    safe_value = value.item()
                else:
                    safe_value = value
                sheet.cell(row=r_idx, column=c_idx, value=safe_value)


# ---------------- Run script ---------------- #

if __name__ == "__main__":
    source_folder = "./data"
    analysis_folder = "./results"

    filter_config = [
        ("CMax", "C_max_tD1_tD2", True, 0.4776),
        ("AUC24h", "AUC_tD1_tD2", True, 0.00796),
        ("Cmin", "C_trough_tD2", False, 0.4776),
    ]

    mic_config = [0.25, 0.5, 1, 2, 4, 8]

    analyzer: PkSimAnalyzer = PkSimAnalyzer(
        source_dir=source_folder,
        output_dir=analysis_folder,
        filter_params=filter_config,
        mic_params=mic_config,
        float_precision=4
    )

    start_time = time.perf_counter()
    analyzer.process_all()

    elapsed = time.perf_counter() - start_time
    print(f"\n✅ Total execution time: {elapsed:.2f} seconds")
